import datetime
import os
import platform
import subprocess
import winsound
from difflib import SequenceMatcher

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter
from unidecode import unidecode

from database.dao.PersonDao import PersonDao
from database.dao.SearchDao import SearchDao
from utils.bcolors import bcolors
from utils.texts import text_error, text_waiting_export_xls, text_export_finished, text_weighted_calculation_performed, \
    text_menu_score, text_menu_score_all, text_menu_score_header, text_menu_score_header_end, text_menu_score_option
from utils.texts import text_error_filter_only_menu
from utils.texts import text_error_score
from utils.texts import text_option_score


class ScoreProfile:
    def __init__(self, database):
        self.zero_ano = 0
        self.um_ano = 12
        self.dois_anos = 24
        self.tres_anos = 36
        self.max_score_technologies = 0
        self.max_score_language = 0
        self.max_score_education = 0
        self.max_score_level = 0
        self.MAX_PERCENT = 80
        self.database = database
        self.TECHNOLOGIES = "technologies"
        self.EXPERIENCE = "experience"
        self.LANGUAGE = "language"
        self.LANGUAGE_LEVEL = "language_level"
        self.FLUENTE_OU_NATIVO = "fluente ou nativo"
        self.AVANCADO = "avançado"
        self.INTERMEDIARIO = "intermediário"
        self.BASICO_A_INTERMEDIARIO = "básico a intermediário"
        self.BASICO = "básico"
        self.LEVEL = "level"
        self.MEDIA = "media"
        self.URL_PROFILE = "url_profile"
        self.EDUCATION = "education"
        self.INGLES = "inglês"
        self.ENGLISH = "english"
        self.MESTRADO = "mestrado"
        self.DOUTORADO = "doutorado"
        self.POS_GRADUACAO = "pós graduação"
        self.POS = "pós"
        self.ESPECIALIZACAO = "especialização"
        self.GRADUACAO = "graduação"
        self.GRADUADO = "graduado"
        self.BACHAREL = "bacharel"
        self.BACHELOR = "bachelor"
        self.BACHARELADO = "bacharelado"
        self.SUPERIOR = "superior"
        self.TECNOLOGO = "tecnólogo"
        self.LEVEL_EDUCATION = "level_education"
        self.SR = "sênior"
        self.PL = "pleno"
        self.JR = "júnior"
        self.UNKNOWN = "Desconhecido"
        self.columns_export = {
            "media": None,
            "url_profile": None,
            "name": None,
            "local": None,
            "education": None,
            "level": None,
            "experience": None,
            "email": None,
            "phone_number": None,
            "language_level": None,
            "technologies": dict()
        }
        self.header_list = ["Media", "URL", "Nome", "Local", "Educação", "Nível", "Experiência", "Email", "Telefone",
                            "Nível de inglês"]
        self.tech_items = ["Tempo", "Certificações", "Selo Linkedin", "Indicações"]
        self.job_characteristics = {
            "technologies": [],
            "language": [self.INGLES, self.ENGLISH],
            "level": [self.SR, self.PL, self.JR]
        }

    def start(self):
        """
        Incialização do cálculo ponderado nos perfis e exportação para XLS.
        """
        try:
            winsound.Beep(250, 100)
            condition_sql = self.__menu()
            winsound.Beep(250, 100)
            option = input(text_option_score)
            self.job_characteristics["technologies"] = [opt.strip() for opt in option.split(",")]
            list_result, search_list = self.__list_person(condition_sql)
            result_list = self.__weighted_calculation(list_result, search_list)
            result = sorted(result_list, key=lambda d: d['scores']['media'], reverse=True)
            self.__export(result)
            print(text_weighted_calculation_performed.format(bcolors.GREEN, bcolors.ENDC))
        except ValueError as e:
            print(text_error_score)
            self.start()

    def __menu(self):
        """
        Inicia o menu.
        """
        try:
            dict_result, text_menu = self.__create_menu()
            option = input(text_menu)
            option = int(option)
            option_dict = dict_result[option]
            return option_dict["condition_sql"]
        except KeyError as e:
            print(text_error_filter_only_menu)
            self.__menu()
        except ValueError as e:
            print(text_error)
            self.__menu()

    def __create_menu(self):
        """
        Faz a criação do menu para selecionar qual pesquisa será feita a exportação para XLS.
        """
        search_list_menu = SearchDao(self.database).select_search_person_group_by_url_filter()
        dict_result = dict()
        string_space = "    "
        for idx, search in enumerate(search_list_menu):
            counter = SearchDao(self.database).search_counter_by_url_filter(search.url_filter)
            option_number = idx + 1
            text = text_menu_score.format(option_number, search.datetime, search.text_filter, counter)
            dict_result[option_number] = {
                "text": text,
                "condition_sql": search.url_filter
            }
        number = len(dict_result.keys())
        number += 1
        counter_all = SearchDao(self.database).search_counter_is_not_null()
        dict_result[number] = {
            "text": text_menu_score_all.format(number, counter_all),
            "condition_sql": None
        }
        text_menu = text_menu_score_header.format(string_space, bcolors.HEADER, bcolors.ENDC)
        for value in dict_result.values():
            text_menu = text_menu + f"{string_space}{value['text'].title()}"
        text_menu = text_menu + text_menu_score_header_end.format(string_space, bcolors.HEADER, bcolors.ENDC)
        text_menu = text_menu + text_menu_score_option.format(string_space, bcolors.BOLD, bcolors.CYAN, bcolors.ENDC,
                                                              bcolors.ENDC)
        return dict_result, text_menu

    def __list_person(self, condition_sql):
        """
        Baseado no filtro que é retornado, traz todos os perfis necessários.
        """
        if condition_sql:
            search_list = SearchDao(self.database).select_search_by_url_filter(condition_sql)
        else:
            search_list = SearchDao(self.database).select_search_person_id_is_not_null()
        list_result = list()
        for search in search_list:
            list_result.append(PersonDao(database=self.database).select_people_by_id(search.person_id))
        return list_result, search_list

    def __export(self, result_list):
        """
        Preparação dos dados para a planilha e exportação do arquivo.
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Profile Scores"
        font_header, border, alignment, pattern_fill, number_format = self.__style_header()
        sheet = self.__create_header_xls(sheet, font_header, border, alignment, pattern_fill)
        columns_export_size = self.__set_size_header()
        row_number = 2
        for row in result_list:
            columns = self.__get_data_for_openxls(row)
            cell_number = 1
            row_number += 1
            for key, value in columns.items():
                if key == self.MEDIA:
                    columns_export_size = self.__set_max_size_col(columns_export_size, key, value)
                    color, msg, hexadecimal = self.__get_color_media(value)
                    cell = sheet.cell(row=row_number, column=cell_number, value=value)
                    cell.font = Font(name='Arial Black', size=11, bold=True, italic=False, vertAlign=None,
                                     underline='none', strike=False, color=hexadecimal, shadow=True)
                    cell.alignment = alignment
                    cell.border = border
                elif key == self.URL_PROFILE:
                    columns_export_size = self.__set_max_size_col(columns_export_size, key, "Acessar")
                    cell = sheet.cell(row=row_number, column=cell_number, value=value)
                    cell.font = Font(size=11, bold=True, italic=False, vertAlign=None,
                                     underline='single', strike=False, color="2986CC", shadow=True)
                    cell.alignment = alignment
                    cell.border = border
                elif key == self.LEVEL:
                    level, font = self.__get_style_or_unknown(value)
                    columns_export_size = self.__set_max_size_col(columns_export_size, key, level)
                    cell = sheet.cell(row=row_number, column=cell_number, value=level)
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = border
                elif key == self.EDUCATION:
                    level, font = self.__get_style_or_unknown(value)
                    columns_export_size = self.__set_max_size_col(columns_export_size, key, level)
                    cell = sheet.cell(row=row_number, column=cell_number, value=level)
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = border
                elif key == self.LANGUAGE_LEVEL:
                    level, font = self.__get_style_or_unknown(value)
                    columns_export_size = self.__set_max_size_col(columns_export_size, key, level)
                    cell = sheet.cell(row=row_number, column=cell_number, value=level)
                    cell.font = font
                    cell.alignment = alignment
                    cell.border = border
                elif key == self.EXPERIENCE and value == self.UNKNOWN:
                    columns_export_size = self.__set_max_size_col(columns_export_size, key, value)
                    cell = sheet.cell(row=row_number, column=cell_number, value=value)
                    cell.font = Font(bold=False, italic=True)
                    cell.alignment = alignment
                    cell.border = border
                elif key == self.TECHNOLOGIES:
                    if columns[self.TECHNOLOGIES]:
                        for key in self.job_characteristics[self.TECHNOLOGIES]:
                            tech_dict = columns[self.TECHNOLOGIES][key]
                            cell_number, tempo = self.__set_cell_tempo(sheet, tech_dict, row_number, cell_number,
                                                                       alignment, border)
                            columns_export_size = self.__set_max_size_col(
                                columns_export_size,
                                f"{unidecode(key.lower())}_{unidecode(self.tech_items[0].lower())}",
                                tempo
                            )
                            cell_number, status_certification = self.__set_cell_certification(sheet, tech_dict,
                                                                                              row_number, cell_number,
                                                                                              alignment, border)
                            columns_export_size = self.__set_max_size_col(
                                columns_export_size,
                                f"{unidecode(key.lower())}_{unidecode(self.tech_items[1].lower())}",
                                status_certification
                            )
                            cell_number, status_verify = self.__set_cell_verify(sheet, tech_dict, row_number,
                                                                                cell_number, alignment, border)
                            columns_export_size = self.__set_max_size_col(
                                columns_export_size,
                                f"{unidecode(key.lower())}_{unidecode(self.tech_items[2].lower())}",
                                status_verify
                            )
                            cell_number, indications = self.__set_cell_indications(sheet, tech_dict, row_number,
                                                                                   cell_number, alignment, border)
                            columns_export_size = self.__set_max_size_col(
                                columns_export_size,
                                f"{unidecode(key.lower())}_{unidecode(self.tech_items[3].lower())}",
                                indications
                            )
                else:
                    columns_export_size = self.__set_max_size_col(columns_export_size, key, value)
                    cell = sheet.cell(row=row_number, column=cell_number, value=value)
                    cell.alignment = alignment
                    cell.border = border
                cell_number += 1
        self.__expand_columns(columns_export_size, sheet)
        name_file = datetime.datetime.now().strftime("%m_%d_%Y_%H%M%S")
        file_path = f"C:\\scrapingLinkedinProfiles\\export\\{name_file}.xlsx"
        print(text_waiting_export_xls)
        workbook.save(file_path)
        print(text_export_finished.format(bcolors.GREEN, bcolors.ENDC, file_path))
        self.__open_file(file_path)
        winsound.MessageBeep()

    def __set_size_header(self):
        """
        Verificar qual o valor máximo de caracter no HEADER e adicionar no dicionário.
        """
        list_tech_header = len(self.job_characteristics["technologies"]) * self.tech_items
        self.header_list.extend(list_tech_header)
        columns_export_size = self.columns_export.copy()
        del columns_export_size["technologies"]
        for value in zip(self.header_list, columns_export_size):
            columns_export_size = self.__set_max_size_col(columns_export_size, value[1], value[0])
        for tech in self.job_characteristics["technologies"]:
            for value in self.tech_items:
                key = f"{unidecode(tech.lower())}_{unidecode(value.lower())}"
                columns_export_size = self.__set_max_size_col(columns_export_size, key, value)
        return columns_export_size

    def __set_max_size_col(self, columns_export_size, key, value, size_cell=4):
        """
        Verificar qual o valor máximo de caracter e adicionar no dicionário somente se o valor for maior do que o existente.
        """
        if key in columns_export_size.keys():
            if columns_export_size[key]:
                columns_export_size[key] = max(columns_export_size[key], len(str(value)) + size_cell)
            else:
                columns_export_size[key] = len(str(value)) + size_cell
        else:
            columns_export_size[key] = len(str(value)) + size_cell
        return columns_export_size

    def __expand_columns(self, columns_export_size, sheet):
        """
        Expandir células da coluna conforme valores recebidos no dicionário.
        """
        for idx, (key, value) in enumerate(columns_export_size.items()):
            if value:
                sheet.column_dimensions[get_column_letter(idx + 1)].width = value

    def __open_file(self, file_path):
        """
        Verifica o SO e abre abre o respectivo programa para XLS.
        """
        if platform.system() == 'Darwin':  # macOS
            subprocess.call(('open', file_path))
        elif platform.system() == 'Windows':  # Windows
            os.startfile(file_path)
        else:  # linux variants
            subprocess.call(('xdg-open', file_path))

    def __get_style_or_unknown(self, value):
        """
        Retorna o texto com a fonte BOLD, caso o valor seja None retorna com a fonte ITALIC e o texto 'Desconhecido'.
        """
        level = value if value else self.UNKNOWN
        font = Font(bold=True, shadow=True) if value else Font(bold=False, italic=True)
        return level, font

    def __set_cell_tempo(self, sheet, tech_dict, row_number, cell_number, alignment, border):
        """
        Prepara a célula da planilha responsável por exibir tempo de experiência na tecnologia.
        """
        tempo, font = self.__get_formatted_experience_time(tech_dict["tempo"])
        cell_tempo = sheet.cell(row=row_number, column=cell_number, value=tempo)
        cell_tempo.alignment = alignment
        cell_tempo.border = border
        cell_tempo.font = font
        cell_number += 1
        return cell_number, tempo

    def __set_cell_certification(self, sheet, tech_dict, row_number, cell_number, alignment, border):
        """
        Prepara a célula da planilha responsável por exibir aa certificações.
        """
        font, pattern_fill = self.__get_style_status(tech_dict["certification"])
        status = self.__get_sim_nao(tech_dict["certification"])
        cell_certification = sheet.cell(row=row_number, column=cell_number, value=status)
        cell_certification.alignment = alignment
        cell_certification.fill = pattern_fill
        cell_certification.border = border
        cell_certification.font = font
        cell_number += 1
        return cell_number, status

    def __set_cell_verify(self, sheet, tech_dict, row_number, cell_number, alignment, border):
        """
        Prepara a célula da planilha responsável por exibir a verificação do Linkedin.
        """
        font, pattern_fill = self.__get_style_status(tech_dict["verify"])
        status = self.__get_sim_nao(tech_dict["verify"])
        cell_verify = sheet.cell(row=row_number, column=cell_number, value=status)
        cell_verify.alignment = alignment
        cell_verify.fill = pattern_fill
        cell_verify.border = border
        cell_verify.font = font
        cell_number += 1
        return cell_number, status

    def __set_cell_indications(self, sheet, tech_dict, row_number, cell_number, alignment, border):
        """
        Retorna SIM caso seja verdadeiro e NÃO caso falso.
        """
        cell_indications = sheet.cell(row=row_number, column=cell_number, value=tech_dict["indications"])
        cell_indications.alignment = alignment
        cell_indications.border = border
        cell_indications.font = Font(bold=True) if tech_dict["indications"] > 0 else Font(bold=False)
        cell_number += 1
        return cell_number, tech_dict["indications"]

    def __get_style_status(self, status):
        """
        Retorna VERDE caso seja verdadeiro e VERMELHO caso falso.
        """
        if status:
            pattern_fill = PatternFill(start_color="00B050", end_color="00b050", fill_type="solid")
            font = Font(bold=True, shadow=True)
            return font, pattern_fill
        else:
            pattern_fill = PatternFill(start_color="FF0000", end_color="ff0000", fill_type="solid")
            font = Font(bold=True, shadow=True)
            return font, pattern_fill

    def __get_sim_nao(self, status):
        """
        Retorna SIM caso seja verdadeiro e NÃO caso falso.
        """
        if status:
            return "SIM"
        else:
            return "NÃO"

    def __get_data_for_openxls(self, row):
        """
        Preparo incial dos dados para a respectiva linha no excel.
        """
        media = row['scores']['media']
        name = row['person'].name
        local = row['person'].local
        url_profile = row['person'].url
        email = row['person'].email
        phone_number = row['person'].phone_number
        language_level = self.__get_is_exist_english_language(row['scores'])
        if language_level:
            language_level = language_level["level"]
        education = None
        for key in row['scores'][self.EDUCATION].keys():
            if key != "level_education":
                education = key.title()
                break
        level = None
        for key, value in row['scores'][self.LEVEL].items():
            if value['score'] > 0:
                level = key.title()
                break
        time_experience = self.__get_time_experience(row['person'].experiences)
        time_experience = time_experience if time_experience else self.UNKNOWN
        self.columns_export["media"] = media
        self.columns_export["url_profile"] = '=HYPERLINK("{}", "{}")'.format(url_profile, "Acessar")
        self.columns_export["name"] = name
        self.columns_export["local"] = local
        self.columns_export["education"] = education
        self.columns_export["level"] = level
        self.columns_export["experience"] = time_experience
        self.columns_export["email"] = email
        self.columns_export["phone_number"] = phone_number
        self.columns_export["language_level"] = language_level
        self.columns_export["technologies"] = row['scores']['technologies']
        return self.columns_export

    def __get_is_exist_english_language(self, data):
        level_list = [data[self.LANGUAGE][self.INGLES]["level"], data[self.LANGUAGE][self.ENGLISH]["level"]]
        language_score = max([data[self.LANGUAGE][self.INGLES]["score"], data[self.LANGUAGE][self.ENGLISH]["score"]])
        if language_score > 0:
            language_level = [i for i in level_list if i is not None]
            if language_level:
                return {"level": language_level[0], "score": language_score}
        return None

    def __get_time_experience(self, experiences):
        """
        Faz o cálculo total do tempo de experiência do perfil e retorna um texto amigável.
        """
        total_years = 0
        total_months = 0
        for experience_list in experiences:
            for experience in experience_list:
                if experience.years:
                    total_years += experience.years
                if experience.months:
                    total_months += experience.months
        months_aux = 0
        years_aux = 0
        for x in range(0, total_months):
            months_aux += 1
            if months_aux == self.um_ano:
                years_aux += 1
                months_aux = 0
        months_remaining = total_months - (years_aux * 12)
        total_years += years_aux
        years_text = self.__get_format_time_experience(total_years, "{} ano", "{} anos")
        months_text = self.__get_format_time_experience(months_remaining, "{} mês", "{} meses")
        return f"{years_text} {months_text}".strip()

    def __get_format_time_experience(self, value, text_singular, text_plural):
        """
        Retorna o texto no plural ou singular dependendo do valor recebido
        """
        result = ""
        if value == 1:
            result = text_singular.format(value)
        elif value > 1:
            result = text_plural.format(value)
        return result

    def __create_header_xls(self, sheet, font_header, border, alignment, pattern_fill):
        """
        Faz as mesclagens necessários no Header e adiciona os estilos.
        """
        item_header = 0
        for item in range(len(self.header_list)):
            item += 1
            item_header += 1
            sheet.merge_cells(start_row=1, start_column=item, end_row=2, end_column=item)

        column_header_num = 0
        for item in self.header_list:
            column_header_num += 1
            cell = sheet.cell(row=1, column=column_header_num, value=item)
            cell.font = font_header
            cell.border = border
            sheet.cell(row=2, column=column_header_num).border = border
            cell.alignment = alignment
            cell.fill = pattern_fill
        # TECH
        for item in range(len(self.job_characteristics["technologies"])):
            item += 1
            item_header += 1
            sheet.merge_cells(start_row=1, start_column=item_header, end_row=1, end_column=item_header + 3)
            item_header += 3

        for item in self.job_characteristics["technologies"]:
            column_header_num += 1
            cell = sheet.cell(row=1, column=column_header_num, value=item.title())
            cell.font = font_header
            cell.border = border
            cell.alignment = alignment
            cell.fill = pattern_fill
            index = column_header_num
            column_header_num += 3
            for tech_item in self.tech_items:
                cell = sheet.cell(row=2, column=index, value=tech_item.title())
                cell.font = font_header
                cell.border = border
                cell.alignment = alignment
                cell.fill = pattern_fill
                index += 1
        return sheet

    def __get_formatted_experience_time(self, tempo):
        """
        Pega o tempo em meses e retorna um texto amigável
        """
        font_bold = Font(bold=True)
        if self.zero_ano < tempo <= self.um_ano:
            return "Menor igual a 1 ano.", font_bold
        elif self.um_ano < tempo <= self.dois_anos:
            return "De 1 a 2 anos.", font_bold
        elif self.dois_anos < tempo <= self.tres_anos:
            return "De 2 a 3 anos.", font_bold
        elif tempo > self.tres_anos:
            return "Maior que 3 anos.", font_bold
        return self.UNKNOWN, Font(bold=False, italic=True)

    def __style_header(self):
        """
        Retorna o estilo do cabeçalho da planilha
        """
        font_header = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None, underline='none',
                           strike=False, color='FF000000')
        side = Side(border_style=None, style='thin', color='000000')
        border = Border(left=side, right=side, top=side, bottom=side, diagonal=side, diagonal_direction=0, outline=side,
                        vertical=side, horizontal=side)
        alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False,
                              shrink_to_fit=False, indent=0)
        pattern_fill = PatternFill(start_color="D6D6D6", end_color="D6D6D6", fill_type="solid")
        number_format = 'General'
        return font_header, border, alignment, pattern_fill, number_format

    def __get_color_media(self, media):
        """
        Baseado na média ponderada é retornado a cor e o texto referente aquele nível de pontos
        """
        hexadecimal = None
        color = None
        msg = None
        if media < 25:
            hexadecimal = "FF0000"
            color = bcolors.RED
            msg = "It's not good..."
        elif 25 <= media < 50:
            hexadecimal = "FFA500"
            color = bcolors.ORANGE
            msg = "Ok..."
        elif 50 <= media <= 75:
            hexadecimal = "E3E300"
            color = bcolors.YELLOW
            msg = "Good!"
        elif media > 75:
            hexadecimal = "008000"
            color = bcolors.GREEN
            msg = "Great!!!"
        return color, msg, hexadecimal

    def __calculate_time(self, years, months):
        """
        Converte os anos em meses e faz a soma do total em meses.
        """
        years = years or 0
        months = months or 0
        result = (years * 12) + months
        return result

    def __weighted_calculation(self, list_result, search_list):
        """
        Planilha: https://docs.google.com/spreadsheets/d/11nilsQIwfTzaWQXd7T3ZyBYDu4JBoX1PD9MRd8S8Ebc/edit?usp=sharing
        ############# Cálculo Ponderado ##############
        |--------------------------------------------|
        |Caracteristicas		  |              Peso|
        |--------------------------------------------|
        |Subtitulo *		                  |  2   |
        |--------------------------------------------|
        |Sobre *		                      |  5   |
        |--------------------------------------------|
        |Sênior		              |              10  |
        |--------------------------------------------|
        |Pleno		              |              7   |
        |--------------------------------------------|
        |Júnior		              |              3   |
        |--------------------------------------------|
        |Inglês	   |    Fluente ou Nativo  |     10  |
        |          |    Avançado	       |     8   |
        |          |    Intermediário	   |     6   |
        |          |    Básico a intermediário | 4   |
        |          |    Básico             |     1   |
        |--------------------------------------------|
        |Tecnologias * |	0 a 1 ano	    |    3   |
        |              |    1 a 2 anos	    |    5   |
        |              |    2 a 3 anos	    |    7   |
        |              |    3 anos +	    |    10  |
        |--------------------------------------------|
        |Skills *	 |    0 a 20 indicações	  |  2   |
        |            |    20 a 40 indicações  |  4   |
        |            |    40 a 60 indicações  |  6   |
        |            |    60 a 80 indicações  |  8   |
        |            |    80 a 99 indicações  |  10  |
        |--------------------------------------------|
        |Selo Linkedin                        |  10  |
        |--------------------------------------------|
        |Certificações *		              |  5   |
        |--------------------------------------------|
        |Mestrado		          |              10  |
        |Doutorado		          |              10  |
        |Pós graduação		      |              8   |
        |Graduação		          |              7   |
        |Técnologo      		  |              5   |
        |--------------------------------------------|

        Média: ( TOTAL_MAX_SCORE_BY_PROFILE / TOTAL_MAX_SCORE_BY_FILTER ) * 100 = MEDIA
        * Ex: ( 75 / 120 ) * 100 = 62,5
        * Ex:
        ##########################################################################
        Tecnologias: Python, React, Node.js, TypeScript, React Native,
        ##########################################################################
        """
        result_list = list()
        for result in zip(list_result, search_list):
            result_dict = dict()
            person = result[0]
            search = result[1]
            score_dict = self.__create_dict_score()
            score_dict = self.__calculation_technologies(person, score_dict)
            score_dict = self.__calculation_language(person, score_dict)
            score_dict = self.__calculation_education(person, score_dict)
            score_dict = self.__calculation_level(person, score_dict)
            result_dict["person"] = person
            result_dict["search"] = search
            result_dict['scores'] = score_dict
            result_dict['scores']["media"] = self.__media(score_dict)
            result_list.append(result_dict)
        return result_list

    def __calculation_technologies(self, person, score_dict):
        """
        Cálculo ponderado das técnologias
        """
        score_dict = self.__set_time_experiences_description(person, score_dict, self.TECHNOLOGIES)
        score_dict, max_score_01 = self.__set_score_skills(person, score_dict, self.TECHNOLOGIES)
        score_dict, max_score_02 = self.__set_score_subtitle(person, score_dict, self.TECHNOLOGIES)
        score_dict, max_score_03 = self.__set_score_about(person, score_dict, self.TECHNOLOGIES)
        score_dict, max_score_04 = self.__set_score_certifications(person, score_dict, self.TECHNOLOGIES)
        score_dict[self.TECHNOLOGIES], max_score_05 = self.__set_score_experiences(score_dict[self.TECHNOLOGIES])
        self.max_score_technologies = sum([max_score_01, max_score_02, max_score_03, max_score_04, max_score_05])
        return score_dict

    def __calculation_language(self, person, score_dict):
        """
        Cálculo ponderado do Inglês
        """
        score_dict, max_score_06 = self.__set_score_languages(person, score_dict, self.LANGUAGE)
        score_dict, max_score_07 = self.__set_score_subtitle(person, score_dict, self.LANGUAGE)
        score_dict, max_score_08 = self.__set_score_about(person, score_dict, self.LANGUAGE)
        score_dict, max_score_09 = self.__set_score_skills(person, score_dict, self.LANGUAGE)
        score_dict, max_score_10 = self.__set_score_certifications(person, score_dict, self.LANGUAGE)
        self.max_score_language = sum([max_score_06, max_score_07, max_score_08, max_score_09, max_score_10])
        return score_dict

    def __calculation_education(self, person, score_dict):
        """
        Cálculo ponderado do nível educacional
        """
        score_dict, max_score_11 = self.__set_score_education(person, score_dict)
        self.max_score_education = sum([max_score_11])
        return score_dict

    def __calculation_level(self, person, score_dict):
        """
        Cálculo ponderado do nível profissional
        """
        score_dict, max_score_12 = self.__set_score_experiences_position(person, score_dict)
        self.max_score_level = sum([max_score_12])
        return score_dict

    def __media(self, score_dict):
        """
        Cálculo da média ponderada
        """
        max_score = self.__max_score(score_dict)
        sum_scores = self.__sum_all_scores(score_dict)
        media = (sum_scores / max_score) * 100
        return round(media, 2)

    def __sum_all_scores(self, score_dict):
        """
        Soma todos os pontos por perfil
        """
        max_score = 0
        for key, value in score_dict.items():
            for k, v in value.items():
                max_score += v["score"]
        return max_score

    def __max_score(self, score_dict):
        """
        Soma todos os pontos máximo de cada chave
        """
        result_media_score = 0
        for key, value in score_dict.items():
            if key == self.TECHNOLOGIES:
                result_media_score += len(value) * self.max_score_technologies
            elif key == self.LANGUAGE:
                result_media_score += len(value) * self.max_score_language
            elif key == self.LEVEL:
                result_media_score += len(value) * self.max_score_level
            elif key == self.EDUCATION:
                result_media_score += len(value) * self.max_score_level
        return result_media_score

    def __create_dict_score(self):
        """
        Cria um dicionario para administrar as pontuações de cada perfil.
        """
        score_dict = dict()
        score_dict["technologies"] = dict()
        for technologie in list(set(self.job_characteristics["technologies"])):
            data_dict = {"certification": False, "verify": False, "indications": 0, "tempo": 0, "score": 0}
            score_dict["technologies"][technologie] = data_dict
        score_dict["language"] = dict()
        for language in list(set(self.job_characteristics["language"])):
            score_dict["language"][language] = {"level": None, "score": 0}
        score_dict["level"] = dict()
        for level in list(set(self.job_characteristics["level"])):
            score_dict["level"][level] = {"score": 0}
        score_dict["education"] = dict()
        score_dict["education"][self.LEVEL_EDUCATION] = {"level": None, "score": 0}
        return score_dict

    def __set_score_subtitle(self, person, score_dict, main_key):
        """
        Adiciona a pontuação do subtitulo baseado nas palavras chaves.
        """
        max_score = 2
        for key, score in score_dict[main_key].items():
            if key in person.subtitle.lower():
                score_dict[main_key][key]['score'] += max_score
        return score_dict, max_score

    def __set_score_about(self, person, score_dict, main_key):
        """
        Adiciona a pontuação ao sobre baseado nas palavras chaves.
        """
        max_score = 5
        for key, score in score_dict[main_key].items():
            if person.about and key in person.about.lower():
                score_dict[main_key][key]['score'] += max_score
        return score_dict, max_score

    def __set_time_experiences_description(self, person, score_dict, main_key):
        """
        Faz a soma do tempo percorrendo as experiências para cada tecnologia.
        """
        for key, score in score_dict[main_key].items():
            for experience_list in person.experiences:
                for experience in experience_list:
                    if key in experience.position.lower() or experience.description and key in experience.description.lower():
                        tempo = self.__calculate_time(experience.years, experience.months)
                        score_dict[main_key][key]['tempo'] += tempo
        return score_dict

    def __set_score_experiences_position(self, person, score_dict):
        """
        Verifica se existe no 'saco de palavras' as palavras 'sênior', 'pleno' e 'júnior'. o loop é interrompido, pois
        só é válido uma única validação buscando a experiência mais recente até a mais antiga.
        """
        max_score = 10
        for experience_list in person.experiences:
            result = self.__get_professional_level(score_dict, experience_list, max_score)
            if result:
                break
        return score_dict, max_score

    def __get_professional_level(self, score_dict, experience_list, max_score=10):
        """
        Caso exista algum nível profissional, é retornado True, se não, Falso.
        """
        for experience in experience_list:
            if unidecode(self.SR) in unidecode(experience.position.lower()) or experience.description and unidecode(
                    self.SR) in unidecode(experience.description.lower()):
                if score_dict[self.LEVEL][self.SR]["score"] == 0:
                    score_dict[self.LEVEL][self.SR]["score"] += max_score
                    return True
            elif self.PL in unidecode(experience.position.lower()) or experience.description and self.PL in unidecode(
                    experience.description.lower()):
                if score_dict[self.LEVEL][self.SR]["score"] == 0:
                    score_dict[self.LEVEL][self.PL]["score"] += 7
                    return True
            elif unidecode(self.JR) in unidecode(experience.position.lower()) or experience.description and unidecode(
                    self.JR) in unidecode(experience.description.lower()):
                if score_dict[self.LEVEL][self.SR]["score"] == 0:
                    score_dict[self.LEVEL][self.JR]["score"] += 3
                    return True
        return False

    def __set_score_experiences(self, technologies):
        """
        Adiciona os pontos de cada tecnologia baseado no tempo de experiência.
        """
        max_score = 10
        for key, value in technologies.items():
            tempo = value["tempo"]
            if self.zero_ano > tempo < self.um_ano:
                technologies[key]['score'] += 3
            elif self.um_ano < tempo < self.dois_anos:
                technologies[key]['score'] += 5
            elif self.dois_anos < tempo < self.tres_anos:
                technologies[key]['score'] += 7
            elif tempo > self.tres_anos:
                technologies[key]['score'] += max_score
        return technologies, max_score

    def __set_score_certifications(self, person, score_dict, main_key):
        """
        Adiciona a pontuação dos certificados.
        """
        max_score = 5
        for key, score in score_dict[main_key].items():
            for certification in person.certifications:
                if key in certification.title.lower():
                    score_dict[main_key][key]['score'] += max_score
                    score_dict[main_key][key]['certification'] = True
                    break
        return score_dict, max_score

    def __set_score_skills(self, person, score_dict, main_key):
        """
        Adiciona a pontuação da skill baseado nas indicações.
        """
        max_score = 10
        for key, score in score_dict[main_key].items():
            for skill in person.skills:
                if key in skill.title.lower():
                    if skill.verify:
                        score_dict[main_key][key]['score'] += max_score
                        score_dict[main_key][key]['verify'] = skill.verify
                    if 0 < skill.indications <= 20:
                        score_dict[main_key][key]['score'] += 2
                        score_dict[main_key][key]['indications'] = skill.indications
                        break
                    elif 20 < skill.indications <= 40:
                        score_dict[main_key][key]['score'] += 4
                        score_dict[main_key][key]['indications'] = skill.indications
                        break
                    elif 40 < skill.indications <= 60:
                        score_dict[main_key][key]['score'] += 6
                        score_dict[main_key][key]['indications'] = skill.indications
                        break
                    elif 60 < skill.indications <= 80:
                        score_dict[main_key][key]['score'] += 8
                        score_dict[main_key][key]['indications'] = skill.indications
                        break
                    elif 80 < skill.indications <= 99:
                        score_dict[main_key][key]['score'] += max_score
                        score_dict[main_key][key]['indications'] = skill.indications
                        break
        return score_dict, (max_score + max_score)

    def __set_score_languages(self, person, score_dict, main_key):
        """
        Adiciona a pontuação do idioma baseado no nível.
        """
        max_score = 10
        for key, score in score_dict[main_key].items():
            for language in person.languages:
                if key in language.language.lower():
                    if language.level:
                        level = language.level.lower().replace("nível", "").strip()
                        if level == self.FLUENTE_OU_NATIVO:
                            score_dict[main_key][key]['score'] += max_score
                            score_dict[main_key][key]['level'] = self.FLUENTE_OU_NATIVO.title()
                        elif level == self.AVANCADO:
                            score_dict[main_key][key]['score'] += 8
                            score_dict[main_key][key]['level'] = self.AVANCADO.title()
                        elif level == self.INTERMEDIARIO:
                            score_dict[main_key][key]['score'] += 6
                            score_dict[main_key][key]['level'] = self.INTERMEDIARIO.title()
                        elif level == self.BASICO_A_INTERMEDIARIO:
                            score_dict[main_key][key]['score'] += 4
                            score_dict[main_key][key]['level'] = self.BASICO_A_INTERMEDIARIO.title()
                        elif level == self.BASICO:
                            score_dict[main_key][key]['score'] += 1
                            score_dict[main_key][key]['level'] = self.BASICO.title()
        return score_dict, max_score

    def __set_score_education(self, person, score_dict):
        """
        Adiciona a pontuação baseado na educação.
        """
        max_score = 10
        for education in person.education:
            if education.level:
                result = {
                    self.MESTRADO: self.__similarity(education.level.lower(), self.MESTRADO),
                    self.DOUTORADO: self.__similarity(education.level.lower(), self.DOUTORADO),
                    self.POS_GRADUACAO: max(self.__similarity(education.level.lower(), self.POS_GRADUACAO),
                                            self.__similarity(education.level.lower(), self.POS),
                                            self.__similarity(education.level.lower(), self.ESPECIALIZACAO)
                                            ),
                    self.GRADUACAO: max(self.__similarity(education.level.lower(), self.GRADUACAO),
                                        self.__similarity(education.level.lower(), self.GRADUADO),
                                        self.__similarity(education.level.lower(), self.BACHAREL),
                                        self.__similarity(education.level.lower(), self.BACHARELADO),
                                        self.__similarity(education.level.lower(), self.BACHELOR),
                                        self.__similarity(education.level.lower(), self.SUPERIOR)
                                        ),
                    self.TECNOLOGO: self.__similarity(education.level.lower(), self.TECNOLOGO)
                }
                key_max_value = max(result, key=result.get)
                all_values = result.values()
                max_value = max(all_values)

                if self.__master_degree(
                        education.level.lower()) or key_max_value == self.MESTRADO and max_value >= self.MAX_PERCENT:
                    score_dict[self.EDUCATION][self.LEVEL_EDUCATION]['score'] += max_score
                    score_dict[self.EDUCATION][self.LEVEL_EDUCATION]['level'] = self.MESTRADO.title()
                    score_dict[self.EDUCATION][self.MESTRADO] = score_dict[self.EDUCATION][self.LEVEL_EDUCATION]
                    del score_dict[self.EDUCATION][self.LEVEL_EDUCATION]
                    break
                elif self.__doctorate_degree(
                        education.level.lower()) or key_max_value == self.DOUTORADO and max_value >= self.MAX_PERCENT:
                    score_dict[self.EDUCATION][self.LEVEL_EDUCATION]['score'] += max_score
                    score_dict[self.EDUCATION][self.LEVEL_EDUCATION]['level'] = self.DOUTORADO.title()
                    score_dict[self.EDUCATION][self.DOUTORADO] = score_dict[self.EDUCATION][self.LEVEL_EDUCATION]
                    del score_dict[self.EDUCATION][self.LEVEL_EDUCATION]
                    break
                elif self.__post_graduate_degree(
                        education.level.lower()) or key_max_value == self.POS_GRADUACAO and max_value >= self.MAX_PERCENT:
                    score_dict[self.EDUCATION][self.LEVEL_EDUCATION]['score'] += 8
                    score_dict[self.EDUCATION][self.LEVEL_EDUCATION][
                        'level'] = f"{self.POS.title()}/{self.ESPECIALIZACAO.title()}"
                    LEVEL = f"{self.POS} ou {self.ESPECIALIZACAO}"
                    score_dict[self.EDUCATION][LEVEL] = score_dict[self.EDUCATION][self.LEVEL_EDUCATION]
                    del score_dict[self.EDUCATION][self.LEVEL_EDUCATION]
                    break
                elif self.__graduate_degree(
                        education.level.lower()) or key_max_value == self.GRADUACAO and max_value >= self.MAX_PERCENT:
                    score_dict[self.EDUCATION][self.LEVEL_EDUCATION]['score'] += 7
                    score_dict[self.EDUCATION][self.LEVEL_EDUCATION]['level'] = self.GRADUACAO.title()
                    score_dict[self.EDUCATION][self.GRADUACAO] = score_dict[self.EDUCATION][self.LEVEL_EDUCATION]
                    del score_dict[self.EDUCATION][self.LEVEL_EDUCATION]
                    break
                elif self.__technologist_degree(
                        education.level.lower()) or key_max_value == self.TECNOLOGO and max_value >= self.MAX_PERCENT:
                    score_dict[self.EDUCATION][self.LEVEL_EDUCATION]['score'] += 5
                    score_dict[self.EDUCATION][self.LEVEL_EDUCATION]['level'] = self.TECNOLOGO.title()
                    score_dict[self.EDUCATION][self.TECNOLOGO] = score_dict[self.EDUCATION][self.LEVEL_EDUCATION]
                    del score_dict[self.EDUCATION][self.LEVEL_EDUCATION]
                    break
        return score_dict, max_score

    def __master_degree(self, level):
        """
        Verifica se tem mestrado na String
        """
        if self.MESTRADO in level:
            return True
        return False

    def __doctorate_degree(self, level):
        """
        Verifica se tem doutorado na String
        """
        if self.DOUTORADO in level:
            return True
        return False

    def __post_graduate_degree(self, level):
        """
        Verifica se tem pós graduação ou especialização na String
        """
        if self.POS_GRADUACAO in level or self.POS in level or self.ESPECIALIZACAO in level:
            return True
        return False

    def __graduate_degree(self, level):
        """
        Verifica se tem graduação na String
        """
        if self.GRADUACAO in level or self.GRADUADO in level or self.BACHAREL in level or self.BACHARELADO in level \
                or self.BACHELOR in level or self.SUPERIOR in level:
            return True
        return False

    def __technologist_degree(self, level):
        """
        Verifica se tem tecnólogo na String
        """
        if self.TECNOLOGO in level:
            return True
        return False

    def __similarity(self, text_db, text_static):
        """
        Verifica se tem string com similaridade e retorna a porcentagem de acerto
        """
        seq = SequenceMatcher(None, text_db, text_static)
        percent = seq.ratio() * 100
        return round(percent, 2)
